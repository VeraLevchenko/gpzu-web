import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.land_passports.constants import XLSX_COLUMNS

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
_DATA_FONT = Font(name="Times New Roman", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def format_area(raw: str) -> str:
    m = re.match(r'^([\d\s.,]+)', raw.strip())
    if m:
        num_str = re.sub(r'[^\d]', '', m.group(1))
        try:
            num = int(num_str)
            return f"{num:,}".replace(",", " ") + " кв. м"
        except ValueError:
            pass
    return raw


def parse_area_numeric(area: str) -> float:
    cleaned = re.sub(r'[^\d.,]', '', area).replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def build_xlsx(rows: list) -> bytes:
    """
    rows — список dict, ключи = названия колонок из XLSX_COLUMNS (кроме «№»).
    Возвращает bytes готового xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Земельные участки"

    border = _thin_border()

    for col_idx, (name, width) in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    for row_num, row_data in enumerate(rows, start=2):
        serial = row_num - 1
        # Колонка 1 — порядковый номер
        cell = ws.cell(row=row_num, column=1, value=serial)
        cell.font = _DATA_FONT
        cell.alignment = _CENTER
        cell.border = border

        for col_idx, (col_name, _) in enumerate(XLSX_COLUMNS[1:], start=2):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = _DATA_FONT
            cell.alignment = _WRAP
            cell.border = border

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
