import logging
import math
import threading
import uuid
import zipfile
from io import BytesIO
from typing import List
from urllib.parse import quote

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook

from api.land_passports.constants import FIXED_VALUES, REQUIRED_COLUMNS, WARN_COLUMNS
from api.land_passports.docx_builder import fill_passport
from api.land_passports.xlsx_builder import build_xlsx, format_area, parse_area_numeric
from core.layers_config import LayerPaths
from models.gp_data import ApplicationInfo, GPData, ParcelInfo, TerritorialZoneInfo
from parsers.egrn_parser import parse_egrn_xml
from parsers.tab_parser import find_ago_for_parcel, parse_ago_layer
from utils.spatial_analysis import perform_spatial_analysis

logger = logging.getLogger("gpzu-web.land-passports")

router = APIRouter(prefix="/api/land-passports", tags=["land-passports"])

# In-memory job store: job_id -> {status, progress, current, total, result?, error?}
_jobs: dict = {}

# Координаты центра Новокузнецкого ГО (СЕВЕР, ВОСТОК) в проекции слоёв
_CITY_CENTER_N = 2_212_262.6
_CITY_CENTER_E = 446_896.1


def _distance_to_center(egrn) -> str:
    """Расстояние от первой точки участка до центра города, км."""
    if not egrn.coordinates:
        return ""
    try:
        c = egrn.coordinates[0]
        x = float(c.x)
        y = float(c.y)
        dist_m = math.sqrt((x - _CITY_CENTER_N) ** 2 + (y - _CITY_CENTER_E) ** 2)
        return f"{dist_m / 1000:.1f} км"
    except (TypeError, ValueError):
        return ""


# ── helpers ──────────────────────────────────────────────────────────────────

def _check_krt(coords: list) -> str:
    """Проверяет, попадает ли участок в слой территорий комплексного развития."""
    if not LayerPaths.KRT.exists():
        logger.warning(f"Слой КРТ не найден: {LayerPaths.KRT}")
        return "Нет"
    try:
        krt_features = parse_ago_layer(LayerPaths.KRT)
        if not krt_features:
            return "Нет"
        result = find_ago_for_parcel(coords, krt_features)
        return "Да" if result is not None else "Нет"
    except Exception as e:
        logger.warning(f"Ошибка проверки КРТ: {e}")
        return "Нет"


def _build_gp_data(egrn) -> GPData:
    coords = [{"num": c.num, "x": c.x, "y": c.y} for c in egrn.coordinates]
    gp = GPData()
    gp.application = ApplicationInfo()
    gp.parcel = ParcelInfo(
        cadnum=egrn.cadnum,
        address=egrn.address,
        area=egrn.area,
        coordinates=coords,
        permitted_use=egrn.permitted_use,
    )
    gp.zone = TerritorialZoneInfo()
    return gp


def _build_row_data(egrn, gp: GPData, krt: str = "Нет", dist_center: str = "") -> dict:
    cadnum = egrn.cadnum or ""
    quarter = ":".join(cadnum.split(":")[:3]) if cadnum else ""
    area_raw = egrn.area or ""
    area_fmt = format_area(area_raw)
    potential = "МКД" if parse_area_numeric(area_raw) > 2500 else "ИЖС"
    zone_code = gp.zone.code or ""
    zone_name = gp.zone.name or ""
    zone_str = f"{zone_code} {zone_name}".strip()
    return {
        "Кадастровый номер": cadnum,
        "Кадастровый квартал": quarter,
        "Субъект РФ": FIXED_VALUES["subject"],
        "Муниципальное образование": FIXED_VALUES["municipality"],
        "Адрес": egrn.address or "",
        "Площадь": area_fmt,
        "Категория земель": egrn.land_category or "",
        "Территориальная зона": zone_str,
        "Вид разрешённого использования": egrn.permitted_use or "",
        "Форма собственности": egrn.ownership_form or "",
        "Кадастровая стоимость": egrn.cadastral_value or "",
        "Потенциал использования": potential,
        "Наличие объектов кап. строительства": "",
        "Наличие инженерных сетей": "",
        "Комплексное развитие территории": krt,
        "Для льготных категорий граждан": FIXED_VALUES["privileged"],
        "Льготная категория": FIXED_VALUES["privileged_category"],
        "Агент АО «ДОМ.РФ»": FIXED_VALUES["dom_rf"],
        "Расстояние до федеральной трассы": FIXED_VALUES["distance_federal"],
        "Расстояние до дороги с твёрдым покрытием": "",
        "Расстояние до центра МО или ГО": dist_center,
        "Расстояние до ближайшего населённого пункта": FIXED_VALUES["distance_settlement"],
        "Инвестиционный портал региона": FIXED_VALUES["investment_portal"],
        "Наименование уполномоченного органа и его контакты": FIXED_VALUES["kumi"],
        "Вовлечён под жилищное строительство": FIXED_VALUES["involved_housing"],
        "Выдан ГПЗУ": "",
        "Выдано разрешение на строительство": FIXED_VALUES["construction_permit"],
        "Отсутствует разрешение на ввод в эксплуатацию": FIXED_VALUES["no_commissioning"],
    }


# ── background workers ────────────────────────────────────────────────────────

def _is_cancelled(job_id: str) -> bool:
    return _jobs.get(job_id, {}).get("status") == "cancelled"


def _run_parse_job(job_id: str, files_data: list):
    total = len(files_data)
    rows, errors = [], []
    for i, (filename, raw) in enumerate(files_data):
        if _is_cancelled(job_id):
            _jobs.pop(job_id, None)
            return
        try:
            egrn = parse_egrn_xml(raw)
            gp = _build_gp_data(egrn)
            coords = [(c["x"], c["y"]) for c in gp.parcel.coordinates] if gp.parcel.coordinates else []
            try:
                gp = perform_spatial_analysis(gp)
            except Exception as e:
                logger.warning(f"Пространственный анализ не выполнен для {filename}: {e}")
            krt = _check_krt(coords)
            dist_center = _distance_to_center(egrn)
            rows.append(_build_row_data(egrn, gp, krt, dist_center))
        except Exception as e:
            logger.error(f"Ошибка парсинга {filename}: {e}")
            errors.append(f"{filename}: {e}")
        _jobs[job_id]["current"] = i + 1
        _jobs[job_id]["progress"] = round((i + 1) / total * 100)

    if _is_cancelled(job_id):
        _jobs.pop(job_id, None)
        return

    if rows:
        _jobs[job_id].update({
            "status": "done",
            "result": build_xlsx(rows),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": "перечень_участков.xlsx",
        })
    else:
        _jobs[job_id].update({
            "status": "error",
            "error": f"Не удалось распарсить ни один файл. Ошибки: {'; '.join(errors)}",
        })


def _run_generate_job(job_id: str, headers: list, data_rows: list):
    total = len(data_rows)
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i, (row_num, row) in enumerate(data_rows):
            if _is_cancelled(job_id):
                _jobs.pop(job_id, None)
                return
            row_data = {
                headers[j]: (str(v) if v is not None else "")
                for j, v in enumerate(row)
                if j < len(headers)
            }
            cadnum = row_data.get("Кадастровый номер", f"row_{row_num}")
            safe_name = cadnum.replace(":", "_").replace("/", "_")
            try:
                docx_bytes = fill_passport(row_data)
                zf.writestr(f"{safe_name}.docx", docx_bytes)
            except Exception as e:
                logger.error(f"Ошибка генерации паспорта для строки {row_num} ({cadnum}): {e}")
            _jobs[job_id]["current"] = i + 1
            _jobs[job_id]["progress"] = round((i + 1) / total * 100)

    if _is_cancelled(job_id):
        _jobs.pop(job_id, None)
        return

    zip_buf.seek(0)
    _jobs[job_id].update({
        "status": "done",
        "result": zip_buf.getvalue(),
        "content_type": "application/zip",
        "filename": "паспорта_участков.zip",
    })


# ── endpoints ─────────────────────────────────────────────────────────────────

@router.post("/parse-egrn")
async def parse_egrn(files: List[UploadFile] = File(...)):
    """Запускает фоновую обработку ЕГРН XML/ZIP → возвращает job_id."""
    files_data = [(upload.filename, await upload.read()) for upload in files]
    job_id = str(uuid.uuid4())
    _jobs[job_id] = {"status": "running", "progress": 0, "current": 0, "total": len(files_data)}
    threading.Thread(target=_run_parse_job, args=(job_id, files_data), daemon=True).start()
    return {"job_id": job_id, "total": len(files_data)}


@router.get("/progress/{job_id}")
async def get_job_progress(job_id: str):
    """Возвращает текущий прогресс задания."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    return {
        "status": job["status"],
        "progress": job["progress"],
        "current": job["current"],
        "total": job["total"],
        "error": job.get("error"),
    }


@router.get("/download/{job_id}")
async def download_job_result(job_id: str):
    """Скачивает результат завершённого задания и удаляет его из памяти."""
    job = _jobs.pop(job_id, None)
    if not job:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    if job["status"] == "error":
        raise HTTPException(status_code=400, detail=job.get("error"))
    if job["status"] != "done":
        raise HTTPException(status_code=400, detail="Обработка ещё не завершена")
    return Response(
        content=job["result"],
        media_type=job["content_type"],
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(job['filename'])}"},
    )


@router.delete("/cancel/{job_id}")
async def cancel_job(job_id: str):
    """Отменяет задание — фоновый поток остановится на следующей итерации."""
    job = _jobs.get(job_id)
    if not job:
        return {"ok": True}  # уже завершено или не существует
    if job["status"] == "running":
        job["status"] = "cancelled"
    return {"ok": True}


@router.post("/validate")
async def validate_xlsx(file: UploadFile = File(...)):
    """
    Проверяет xlsx на наличие обязательных колонок и пустых ячеек в них.
    400 — если отсутствуют колонки; 200 — с предупреждениями о пустых ячейках.
    """
    raw = await file.read()
    wb = load_workbook(BytesIO(raw), read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail={"missing_columns": REQUIRED_COLUMNS})

    headers = [str(h).strip() if h else "" for h in header_row]
    col_index = {name: idx for idx, name in enumerate(headers)}

    missing = [col for col in REQUIRED_COLUMNS if col not in col_index]
    if missing:
        raise HTTPException(status_code=400, detail={"missing_columns": missing})

    warnings = []
    for row_num, row in enumerate(rows_iter, start=2):
        for col in WARN_COLUMNS:
            if col not in col_index:
                continue
            idx = col_index[col]
            val = row[idx] if idx < len(row) else None
            if not val or str(val).strip() == "":
                warnings.append(f"Строка {row_num}: пусто поле «{col}»")

    wb.close()
    return {"ok": True, "warnings": warnings}


@router.post("/generate")
async def generate_passports(file: UploadFile = File(...)):
    """Запускает фоновую генерацию docx-паспортов → возвращает job_id."""
    raw = await file.read()
    wb = load_workbook(BytesIO(raw), read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Файл пустой")

    headers = [str(h).strip() if h else "" for h in header_row]
    data_rows = list(enumerate(rows_iter, start=2))
    wb.close()

    if not data_rows:
        raise HTTPException(status_code=400, detail="Файл не содержит данных")

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {"status": "running", "progress": 0, "current": 0, "total": len(data_rows)}
    threading.Thread(target=_run_generate_job, args=(job_id, headers, data_rows), daemon=True).start()
    return {"job_id": job_id, "total": len(data_rows)}
