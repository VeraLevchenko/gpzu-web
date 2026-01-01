# generator/tu_requests_builder.py
"""
Генератор запросов технических условий в РСО.

ОБНОВЛЕНО (01.01.2025):
- ✅ Запись в базу данных PostgreSQL
- ✅ Автоматическое создание Application если не существует
- ✅ Создание записей TuRequest для каждого РСО
- ✅ Сохранение файлов как вложений
- ✅ Дублирование в Excel (для переходного периода)
- ✅ Защита от дублирования запросов в одно РСО
"""

from __future__ import annotations
from io import BytesIO
from pathlib import Path
from typing import List, Tuple, Dict, Optional
from datetime import datetime
import logging

from docxtpl import DocxTemplate
from openpyxl import load_workbook
from filelock import FileLock, Timeout

# ================ НАСТРОЙКИ ================ #

BASE_DIR = Path(__file__).resolve().parents[1]
TU_TEMPLATES_DIR = BASE_DIR / "templates" / "tu"
TU_JOURNAL_PATH = BASE_DIR / "Журнал_регистрации_ТУ_ГПЗУ.xlsx"
TU_JOURNAL_LOCK_PATH = BASE_DIR / "Журнал_регистрации_ТУ_ГПЗУ.xlsx.lock"
JOURNAL_SHEET_NAME = "Лист1"

# Директория для вложений
ATTACHMENTS_DIR = BASE_DIR / "uploads" / "attachments" / "tu"
ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)

# Конфигурация РСО: (код, название_для_журнала, путь_к_шаблону)
TEMPLATE_CONFIG = [
    ("vodokanal", "ООО «Водоканал»", TU_TEMPLATES_DIR / "Водоканал.docx"),
    ("gaz", "филиал ООО «Газпром газораспределение Сибирь»", TU_TEMPLATES_DIR / "Газоснабжение.docx"),
    ("teplo", "ООО «ЭнергоТранзит», ООО «Новокузнецкая теплосетевая компания»", TU_TEMPLATES_DIR / "Теплоснабжение.docx"),
]

# Маппинг кодов на русские названия для файлов
RSO_FILE_NAMES = {
    "vodokanal": "Водоканал",
    "gaz": "Газоснабжение",
    "teplo": "Теплоснабжение",
}

logger = logging.getLogger("gpzu-web.tu_requests_builder")


# ================ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ================ #

def _format_area(area: Optional[str]) -> str:
    """Форматирует площадь (убирает .0 в конце)"""
    if not area:
        return ""
    s = str(area).strip().replace(",", ".")
    if s.endswith(".0"):
        s = s[:-2]
    return s


def convert_date_format(date_str: str) -> str:
    """
    Конвертирует дату из формата «08» декабря 2025 г. в 08.12.2025
    """
    if not date_str:
        return "—"
    
    # Если уже в нужном формате (DD.MM.YYYY)
    if "." in date_str and len(date_str.split(".")) == 3:
        return date_str
    
    # Словарь месяцев
    months = {
        "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
        "мая": "05", "июня": "06", "июля": "07", "августа": "08",
        "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12",
    }
    
    try:
        day_part = date_str.split("«", 1)[1].split("»", 1)[0].strip()
        day = day_part.zfill(2)
        
        rest = date_str.split("»", 1)[1].strip()
        rest = rest.replace("г.", "").replace("г", "").strip()
        parts = rest.split()
        
        if len(parts) >= 2:
            month_name = parts[0].lower()
            year = parts[1]
            month_num = months.get(month_name)
            
            if month_num and year:
                return f"{day}.{month_num}.{year}"
    
    except Exception as ex:
        logger.warning(f"⚠️ Не удалось конвертировать дату '{date_str}': {ex}")
    
    return date_str


def build_tu_context(
    cadnum: str,
    address: str,
    area: str,
    vri: str,
    app_number: str,
    app_date: str,
    out_num: str,
    out_date: str
) -> Dict[str, str]:
    """Формирует контекст для шаблона"""
    return {
        "APP_NUMBER": app_number or "",
        "APP_DATE": app_date or "",
        "CADNUM": cadnum or "",
        "AREA": _format_area(area),
        "VRI": vri or "",
        "ADDRESS": address or "",
        "OUT_NUM": out_num or "",
        "OUT_DATE": out_date or "",
    }


def _render_doc(template_path: Path, context: Dict[str, str]) -> bytes:
    """Рендерит документ из шаблона"""
    tpl = DocxTemplate(str(template_path))
    tpl.render(context)
    bio = BytesIO()
    tpl.save(bio)
    return bio.getvalue()


def get_or_create_application(
    app_number: str,
    app_date: str,
    applicant: str,
    cadnum: str,
    address: str,
    area: str,
    vri: str,
    db_session
) -> int:
    """
    Находит существующее заявление или создает новое.
    
    Returns:
        ID заявления (application_id)
    """
    from models.application import Application
    
    # Ищем существующее заявление
    existing = db_session.query(Application).filter(Application.number == app_number).first()
    
    if existing:
        logger.info(f"✅ Найдено существующее заявление #{app_number} (ID: {existing.id})")
        return existing.id
    
    # Создаем новое заявление
    logger.info(f"📝 Создаем новое заявление #{app_number}")
    
    application = Application(
        number=app_number,
        date=convert_date_format(app_date),
        applicant=applicant or "—",
        phone="—",  # Для ТУ телефон не передается, ставим заглушку
        email="—",  # Для ТУ email не передается, ставим заглушку
        cadnum=cadnum,
        address=address,
        area=float(area) if area else None,
        permitted_use=vri,
        status='in_progress'
    )
    
    db_session.add(application)
    db_session.flush()
    
    logger.info(f"✅ Заявление создано (ID: {application.id})")
    return application.id


def save_tu_to_database(
    application_id: int,
    rso_code: str,
    rso_name: str,
    out_number: int,
    out_date: str,
    attachment_path: str,
    db_session
) -> Optional[int]:
    """
    Сохраняет запрос ТУ в базу данных.
    
    Returns:
        ID созданной записи TuRequest или None при ошибке
    """
    try:
        from models.tu_request import TuRequest
        
        # Проверяем нет ли уже запроса в это РСО
        existing = db_session.query(TuRequest).filter(
            TuRequest.application_id == application_id,
            TuRequest.rso_type == rso_code
        ).first()
        
        if existing:
            logger.warning(f"⚠️ Запрос в {rso_name} уже существует (ID: {existing.id})")
            return existing.id
        
        # Извлекаем год из даты
        try:
            year = int(out_date.split('.')[-1])
        except:
            year = datetime.now().year
        
        # Создаем запись
        tu_request = TuRequest(
            application_id=application_id,
            out_number=out_number,
            out_date=out_date,
            out_year=year,
            rso_type=rso_code,
            rso_name=rso_name,
            attachment=attachment_path,
        )
        
        db_session.add(tu_request)
        db_session.flush()
        
        logger.info(f"✅ Запрос ТУ в {rso_name} сохранен в БД (ID: {tu_request.id})")
        return tu_request.id
        
    except Exception as ex:
        logger.error(f"❌ Ошибка сохранения ТУ в БД: {ex}")
        return None


def write_tu_to_excel_journal(
    app_number: str,
    app_date: str,
    applicant: str,
    cadnum: str,
    address: str,
    rso_name: str,
    out_number: int,
    out_date: str
) -> bool:
    """
    Дублирует запись в Excel журнал (для переходного периода).
    """
    if not TU_JOURNAL_PATH.exists():
        logger.info("ℹ️ Excel журнал не найден, пропускаем дублирование")
        return False
    
    lock = FileLock(str(TU_JOURNAL_LOCK_PATH), timeout=10)
    
    try:
        with lock:
            wb = load_workbook(TU_JOURNAL_PATH)
            ws = wb[JOURNAL_SHEET_NAME]
            
            headers = {cell.value: cell.column for cell in ws[1] if cell.value}
            
            col_out_num = headers.get("Исходящий номер")
            col_out_date = headers.get("Исходящая дата")
            col_app_num = headers.get("Номер заявления")
            col_app_date = headers.get("Дата заявления")
            col_applicant = headers.get("Заявитель")
            col_cadnum = headers.get("Кадастровый номер земельного участка")
            col_address = headers.get("Адрес")
            col_rso = headers.get("РСО")
            
            new_row = ws.max_row + 1
            if col_out_num: ws.cell(row=new_row, column=col_out_num, value=out_number)
            if col_out_date: ws.cell(row=new_row, column=col_out_date, value=out_date)
            if col_app_num: ws.cell(row=new_row, column=col_app_num, value=app_number)
            if col_app_date: ws.cell(row=new_row, column=col_app_date, value=app_date)
            if col_applicant: ws.cell(row=new_row, column=col_applicant, value=applicant)
            if col_cadnum: ws.cell(row=new_row, column=col_cadnum, value=cadnum)
            if col_address: ws.cell(row=new_row, column=col_address, value=address)
            if col_rso: ws.cell(row=new_row, column=col_rso, value=rso_name)
            
            wb.save(TU_JOURNAL_PATH)
            logger.info(f"✅ Запрос ТУ №{out_number} в {rso_name} продублирован в Excel")
            return True
            
    except Exception as ex:
        logger.warning(f"⚠️ Не удалось записать в Excel: {ex}")
        return False


# ================ ОСНОВНАЯ ФУНКЦИЯ ================ #

def build_tu_docs_with_outgoing(
    cadnum: str,
    address: str,
    area: str,
    vri: str,
    app_number: str,
    app_date: str,
    applicant: str
) -> List[Tuple[str, bytes]]:
    """
    Формирует запросы ТУ во все РСО с регистрацией в БД.
    
    ОБНОВЛЕНО (01.01.2025): Теперь записывает в БД + создает Application + сохраняет файлы.
    
    Args:
        cadnum: Кадастровый номер
        address: Адрес участка
        area: Площадь участка
        vri: Вид разрешенного использования
        app_number: Номер заявления
        app_date: Дата заявления
        applicant: Заявитель
    
    Returns:
        List[(filename, docx_bytes)] - список файлов
    
    Raises:
        RuntimeError: Проблемы с генерацией
    """
    
    logger.info(f"📝 Генерация запросов ТУ для заявления {app_number}")
    
    from database import SessionLocal
    from models.tu_request import get_next_tu_number
    
    db = SessionLocal()
    docs: List[Tuple[str, bytes]] = []
    
    try:
        # ========== ШАГ 1: ПОЛУЧАЕМ ИЛИ СОЗДАЕМ APPLICATION ========== #
        
        application_id = get_or_create_application(
            app_number=app_number,
            app_date=app_date,
            applicant=applicant,
            cadnum=cadnum,
            address=address,
            area=area,
            vri=vri,
            db_session=db
        )
        
        # ========== ШАГ 2: ГЕНЕРИРУЕМ ДОКУМЕНТЫ ДЛЯ КАЖДОГО РСО ========== #
        
        current_year = datetime.now().year
        today_str = datetime.now().strftime("%d.%m.%Y")
        cad_for_filename = cadnum.replace(":", " ")
        
        for rso_code, rso_name, tpl_path in TEMPLATE_CONFIG:
            if not tpl_path.exists():
                logger.warning(f"⚠️ Шаблон не найден: {tpl_path}")
                continue
            
            # Получаем следующий номер из БД
            out_number = get_next_tu_number(db, year=current_year)
            out_date_str = today_str
            
            logger.info(f"📋 Генерируем запрос в {rso_name}, исх. №{out_number}")
            
            # Формируем документ
            ctx = build_tu_context(
                cadnum=cadnum,
                address=address,
                area=area,
                vri=vri,
                app_number=app_number,
                app_date=app_date,
                out_num=str(out_number),
                out_date=out_date_str
            )
            
            content = _render_doc(tpl_path, ctx)
            
            # Сохраняем файл
            file_suffix = RSO_FILE_NAMES.get(rso_code, rso_code)
            filename = f"ТУ_{file_suffix}_{cad_for_filename}.docx"
            file_path = ATTACHMENTS_DIR / filename
            
            with open(file_path, "wb") as f:
                f.write(content)
            
            logger.info(f"💾 Файл сохранен: {file_path}")
            
            # Записываем в БД
            tu_id = save_tu_to_database(
                application_id=application_id,
                rso_code=rso_code,
                rso_name=rso_name,
                out_number=out_number,
                out_date=out_date_str,
                attachment_path=str(file_path),
                db_session=db
            )
            
            if tu_id:
                logger.info(f"✅ Запись ТУ создана (ID: {tu_id})")
            
            # Дублируем в Excel
            write_tu_to_excel_journal(
                app_number=app_number,
                app_date=app_date,
                applicant=applicant,
                cadnum=cadnum,
                address=address,
                rso_name=rso_name,
                out_number=out_number,
                out_date=out_date_str
            )
            
            # Добавляем в результат
            docs.append((filename, content))
        
        # Коммитим все изменения
        db.commit()
        
        logger.info(f"✅ Сгенерировано {len(docs)} запросов ТУ")
        
        return docs
        
    except Exception as ex:
        db.rollback()
        raise RuntimeError(f"Ошибка генерации запросов ТУ: {ex}")
    finally:
        db.close()